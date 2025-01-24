import os
import pandas as pd
from openpyxl import load_workbook
import re

# Specify directories
footing_directory = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 
                                 'cpt_based_tool', 'gpt')
output_directory = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 
                                'cpt_based_tool', 'gpt')
os.makedirs(output_directory, exist_ok=True)

# Read template file
cpt_lcpc_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 
                             'cpt_based_tool', 'CPT_lcpc.xlsx')

# Get all footing_ files
file_paths = {
    filename: os.path.join(footing_directory, filename)
    for filename in os.listdir(footing_directory)
    if filename.startswith('footing_') and filename.endswith('.xlsx')
}

for name, path in file_paths.items():
    try:
        # Read data from "CPT data & Bearing Capacity" sheet
        footing_data = pd.read_excel(path, sheet_name="CPT data & Bearing Capacity")
        footing_columns = footing_data.iloc[1:, :4]  # Get first 4 columns

        # Load the template workbook
        workbook = load_workbook(cpt_lcpc_path)
        sheet = workbook["CPT data reduction"]
        template_max_row = sheet.max_row

        # Find last row with formulas
        formula_row = None
        for row in range(template_max_row, 0, -1):
            cell = sheet.cell(row=row, column=5)  # Check column E
            if cell.data_type == 'f':
                formula_row = row
                break
                
        print(f"Found formulas in row {formula_row}")
        
        # Store formulas from the found row
        last_row_formulas = {}
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=formula_row, column=col)
            last_row_formulas[col] = {
                'value': cell.value,
                'data_type': cell.data_type
            }
            print(f"Col {col}: {cell.value}")

        # 打印调试信息
        print(f"Template max row: {template_max_row}")
        print(f"Number of formulas collected: {len(last_row_formulas)}")
        for col, info in last_row_formulas.items():
            print(f"Column {col}: value={info['value']}, type={info['data_type']}")

        
        # Store formulas from the template's last row (修改后的版本)
        last_row_formulas = {}
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=template_max_row, column=col)
            # 打印每列的公式信息
            print(f"Column {col}: value={cell.value}, type={cell.data_type}")
            last_row_formulas[col] = {
                'value': cell.value,
                'data_type': cell.data_type
            }

        def update_formula(formula, new_row, template_row):
            """Update formula to adjust row references based on the new row."""
            if not formula or not isinstance(formula, str):
                return formula
                
            def replace_row(match):
                col, row = match.group(1), int(match.group(2))
                if '$' in match.group(0):
                    return match.group(0)
                relative_pos = row - template_row
                new_row_num = new_row + relative_pos
                return f"{col}{new_row_num}"
                
            # 修改正则表达式以匹配更复杂的公式
            updated = re.sub(r'([A-Z]+)(\d+)', replace_row, formula)
            print(f"Original formula: {formula}")
            print(f"Updated formula: {updated}")
            return updated

        # Copy formulas to new rows
        for row_idx, row_data in enumerate(footing_columns.itertuples(index=False), start=3):
            # Input data (columns A-D)
            for col_idx, value in enumerate(row_data, start=1):
                sheet.cell(row=row_idx, column=col_idx).value = value

            # Copy formulas for columns E and beyond
            if row_idx > template_max_row:
                for col in range(5, sheet.max_column + 1):
                    if col in last_row_formulas:
                        formula_info = last_row_formulas[col]
                        if formula_info['data_type'] == 'f':
                            new_formula = update_formula(formula_info['value'], row_idx, formula_row)
                            cell = sheet.cell(row=row_idx, column=col)
                            cell.value = new_formula
              
        # Debug print to verify formula extension
        print(f"\nVerifying formulas after row {template_max_row}:")
        for row in range(template_max_row + 1, template_max_row + 3):
            for col in range(5, 8):  # Check first few columns after D
                cell = sheet.cell(row=row, column=col)
                print(f"Row {row}, Col {col}: {cell.value}")

        # Save the updated file
        output_path = os.path.join(output_directory, f'lcpc_{name}')
        workbook.save(output_path)
        print(f"File saved: {output_path}")
        
    except Exception as e:
        print(f"Error processing file {name}: {e}")
