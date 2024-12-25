import os
import pandas as pd
from openpyxl import load_workbook

# Specify directories
reduce_directory = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'cpt_based_tool', 'gpt')
output_directory = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'cpt_based_tool', 'gpt')
os.makedirs(output_directory, exist_ok=True)

# Read template file
cpt_footing_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'cpt_based_tool', 'CPT_footing.xlsx')

# Get all reduced_ files
file_paths = {}
for filename in os.listdir(reduce_directory):
    if filename.startswith('reduced_') and filename.endswith('.xlsx'):
        file_paths[filename] = os.path.join(reduce_directory, filename)

for name, path in file_paths.items():
    try:
        # Read reduced data
        reduced_data = pd.read_excel(path)
        reduced_columns = reduced_data.iloc[0:, :4]  # Data starts from row 2 (index 1)
        gwl_value = reduced_data.iloc[0, 4]  # GWL value is at E2 (row 2, column 5)

        # Load the template workbook
        workbook = load_workbook(cpt_footing_path)

        # Access the specific sheet
        sheet = workbook["CPT data & Bearing Capacity"]

        # Update GWL value in the template
        sheet.cell(row=1, column=2).value = gwl_value  # B1

        # Replace the first four columns in the sheet
        for row_idx, row_data in enumerate(reduced_columns.itertuples(index=False), start=3):  # Template data starts at row 3
            for col_idx, value in enumerate(row_data, start=1):
                sheet.cell(row=row_idx, column=col_idx).value = value

        # Clear remaining rows if reduced data has fewer rows
        for row_idx in range(len(reduced_columns) + 3, sheet.max_row + 1):
            for col_idx in range(1, 5):
                sheet.cell(row=row_idx, column=col_idx).value = None

        # Save the updated workbook
        output_path = os.path.join(output_directory, f'updated_{name}')
        workbook.save(output_path)

    except Exception as e:
        print(f"Error processing file {name}: {e}")
